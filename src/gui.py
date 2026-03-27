__author__ = "Süleyman Bozkurt"
__version__ = "v3.0.0"
__maintainer__ = "Süleyman Bozkurt"
__email__ = "sbozkurt.mbg@gmail.com"
__date__ = '18.01.2022'
__update__ = '26.03.2026'

import os
import sys
import re
import random
from threading import Thread
from tkinter import (
    Tk, Frame, Label, Button, StringVar, Radiobutton, END,
    LEFT, W, E, N, S, DISABLED, NORMAL, BOTH, X, Y, TOP, BOTTOM, RIGHT,
    LabelFrame, WORD, HORIZONTAL, Text, CENTER
)
from tkinter.font import Font
from tkinter.scrolledtext import ScrolledText
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook
import openpyxl.styles as sty
from datetime import datetime
import pandas as pd

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from src.functions import mePROD

# ── Colour Palette ──────────────────────────────────────────────
BG           = '#E8ECF1'      # main window background (cool gray)
CARD_BG      = '#FFFFFF'      # card / section background
HEADER_BG    = '#1B2A4A'      # dark navy header
HEADER_FG    = '#FFFFFF'
SECTION_HDR  = '#E2E8F0'      # unified section header strip

ACCENT       = '#4F6D9B'      # muted steel-blue for all buttons
ACCENT_HOVER = '#3B5578'

TEXT_PRIMARY  = '#1E293B'      # dark text
TEXT_SECOND   = '#64748B'      # muted text
BORDER        = '#CBD5E1'      # subtle border

STATUS_BG    = '#1E1E2E'      # dark console background
STATUS_FG    = '#A5F3C4'      # green console text


def _hover_bind(widget, normal_bg, hover_bg):
    """Attach enter/leave colour swap to a widget."""
    widget.bind('<Enter>', lambda e: widget.configure(bg=hover_bg))
    widget.bind('<Leave>', lambda e: widget.configure(bg=normal_bg))


class MyWindow:
    def __init__(self, parent):
        self.filename_condition = ''
        self.root = parent
        parent.configure(bg=BG)

        # ── Fonts ───────────────────────────────────────────────
        self.f_header   = Font(family='Segoe UI', size=14, weight='bold')
        self.f_sub      = Font(family='Segoe UI', size=9)
        self.f_section  = Font(family='Segoe UI', size=9, weight='bold')
        self.f_label    = Font(family='Segoe UI', size=9)
        self.f_btn      = Font(family='Segoe UI', size=9, weight='bold')
        self.f_btn_lg   = Font(family='Segoe UI', size=9, weight='bold')
        self.f_status   = Font(family='Consolas', size=9)
        self.f_file     = Font(family='Segoe UI', size=9, slant='italic')
        self.f_entry    = Font(family='Segoe UI', size=9)

        # ── ttk Style for radio buttons (default theme) ────────
        style = ttk.Style()
        style.configure('Card.TRadiobutton',
                        font=('Segoe UI', 9),
                        background=CARD_BG,
                        foreground=TEXT_PRIMARY,
                        focuscolor='',
                        indicatorsize=14,
                        padding=(4, 2))
        style.map('Card.TRadiobutton',
                  background=[('active', CARD_BG)],
                  indicatorcolor=[('selected', ACCENT), ('!selected', '#CBD5E1')])

        # ── Header bar ─────────────────────────────────────────
        header = Frame(parent, bg=HEADER_BG, height=48)
        header.pack(fill=X, side=TOP)
        header.pack_propagate(False)

        Label(header, text='mePROD App', font=self.f_header,
              bg=HEADER_BG, fg=HEADER_FG).pack(side=LEFT, padx=16, pady=8)
        Label(header, text=f'{__version__}  |  DynaTMT 2.9.4  |  PBLMM 2.1.3',
              font=self.f_sub, bg=HEADER_BG, fg='#94A3B8').pack(side=LEFT, padx=4)
        Label(header, text='S. Bozkurt @2026', font=self.f_sub,
              bg=HEADER_BG, fg='#94A3B8').pack(side=RIGHT, padx=16)

        # ── Body ────────────────────────────────────────────────
        body = Frame(parent, bg=BG, padx=14, pady=6)
        body.pack(fill=BOTH, expand=True)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(4, weight=1)   # console row expands

        # ── Card helper ─────────────────────────────────────────
        def card(parent_frame, row, col, title, colspan=1, sticky='ew', **kw):
            outer = Frame(parent_frame, bg=BG)
            outer.grid(row=row, column=col, columnspan=colspan,
                       sticky=sticky, padx=3, pady=3, **kw)
            outer.columnconfigure(0, weight=1)
            hdr = Frame(outer, bg=SECTION_HDR, highlightbackground=BORDER,
                        highlightthickness=1)
            hdr.pack(fill=X)
            Label(hdr, text=title, font=self.f_section, bg=SECTION_HDR,
                  fg=TEXT_PRIMARY, anchor=W, padx=10, pady=3).pack(fill=X)
            content = Frame(outer, bg=CARD_BG, highlightbackground=BORDER,
                            highlightthickness=1, padx=10, pady=6)
            content.pack(fill=BOTH, expand=True)
            content.columnconfigure(0, weight=1)
            return content

        # ═══════════════════════════════════════════════════════
        #  ROW 0 — MS Level
        # ═══════════════════════════════════════════════════════
        ms_c = card(body, 0, 0, '  MS Acquisition Level', colspan=2)

        self.msLevelVar = StringVar(value='MS2')
        rb_frame = Frame(ms_c, bg=CARD_BG)
        rb_frame.pack(anchor=W)
        for txt, val in [('MS2  (IT adjustment + baseline correction)', 'MS2'),
                         ('MS3  (no IT adj., no baseline correction)', 'MS3')]:
            ttk.Radiobutton(rb_frame, text=txt, value=val, variable=self.msLevelVar,
                            style='Card.TRadiobutton').pack(side=LEFT, padx=(0, 28))

        # ═══════════════════════════════════════════════════════
        #  ROW 1 — Normalization + Statistics side-by-side
        # ═══════════════════════════════════════════════════════
        row1 = Frame(body, bg=BG)
        row1.grid(row=1, column=0, columnspan=2, sticky='ew')
        row1.columnconfigure(0, weight=1)
        row1.columnconfigure(1, weight=1)

        norm_c = card(row1, 0, 0, '  Normalization')
        self.normVar = StringVar(value='total')
        nf = Frame(norm_c, bg=CARD_BG)
        nf.pack(anchor=W)
        for txt, val in [('Total Intensity', 'total'), ('Median', 'median'), ('TMM', 'TMM')]:
            ttk.Radiobutton(nf, text=txt, value=val, variable=self.normVar,
                            style='Card.TRadiobutton').pack(side=LEFT, padx=(0, 14))

        stat_c = card(row1, 0, 1, '  Statistical Method')
        self.statisticVar = StringVar(value='LMM')
        sf = Frame(stat_c, bg=CARD_BG)
        sf.pack(anchor=W)
        for txt, val in [('Linear Mixed Model', 'LMM'), ('Unpaired t-test', 'ttest')]:
            ttk.Radiobutton(sf, text=txt, value=val, variable=self.statisticVar,
                            style='Card.TRadiobutton').pack(side=LEFT, padx=(0, 14))

        # ═══════════════════════════════════════════════════════
        #  ROW 2 — Input (left-aligned, full width)
        # ═══════════════════════════════════════════════════════
        inp_c = card(body, 2, 0, '  Input', colspan=2)

        # PSMs row — label, browse, filename all left-to-right
        psm_row = Frame(inp_c, bg=CARD_BG)
        psm_row.pack(fill=X, pady=(0, 4))

        Label(psm_row, text='PSMs File:', font=self.f_label, bg=CARD_BG,
              fg=TEXT_PRIMARY,
              width=11, anchor=W).pack(side=LEFT, padx=(0, 8)) 

        self.browseButton = Button(psm_row, text=' Browse... ', font=self.f_btn,
                                   bg=ACCENT, fg='white', bd=0, padx=10, pady=2,
                                   activebackground=ACCENT_HOVER, activeforeground='white',
                                   cursor='hand2', command=self.browse,
                                   relief='flat', highlightthickness=0)
        self.browseButton.pack(side=LEFT, padx=(0, 10))
        _hover_bind(self.browseButton, ACCENT, ACCENT_HOVER)

        self.fileLabel = Label(psm_row, text='No file selected', font=self.f_file,
                               bg=CARD_BG, fg=TEXT_SECOND, anchor=W)
        self.fileLabel.pack(side=LEFT, fill=X, expand=True)

        # Output name row — label then entry, left-to-right
        out_row = Frame(inp_c, bg=CARD_BG)
        out_row.pack(fill=X, pady=(0, 2))

        Label(out_row, text='Output Name:', font=self.f_label, bg=CARD_BG,
              fg=TEXT_PRIMARY).pack(side=LEFT, padx=(0, 8))

        self.outputNamebox = Text(out_row, font=self.f_entry, bd=0, height=1,
                                  highlightthickness=1, highlightbackground=BORDER,
                                  highlightcolor=ACCENT, bg='#F8FAFC',
                                  insertbackground=TEXT_PRIMARY, wrap=WORD)
        self.outputNamebox.pack(side=LEFT, fill=X, expand=True, ipady=3)

        # ═══════════════════════════════════════════════════════
        #  ROW 3 — Conditions + Pairs
        # ═══════════════════════════════════════════════════════
        row3 = Frame(body, bg=BG)
        row3.grid(row=3, column=0, columnspan=2, sticky='ew')
        row3.columnconfigure(0, weight=3)
        row3.columnconfigure(1, weight=2)

        cond_c = card(row3, 0, 0, '  Conditions')
        self.conditionbox = Text(cond_c, font=self.f_entry, bd=0, height=3, wrap=WORD,
                                 highlightthickness=1, highlightbackground=BORDER,
                                 highlightcolor=ACCENT, bg='#F8FAFC',
                                 insertbackground=TEXT_PRIMARY)
        self.conditionbox.pack(fill=X, pady=(0, 6))

        condtionsFromText = open('condtions.txt').read()
        self.conditionbox.insert(END, condtionsFromText)

        self.browseButtonCondition = Button(cond_c, text=' Browse... ', font=self.f_btn,
                                            bg=ACCENT, fg='white', bd=0, padx=10, pady=2,
                                            activebackground=ACCENT_HOVER, activeforeground='white',
                                            cursor='hand2', command=self.browse_condition,
                                            relief='flat', highlightthickness=0)
        self.browseButtonCondition.pack(anchor=W)
        _hover_bind(self.browseButtonCondition, ACCENT, ACCENT_HOVER)

        pair_c = card(row3, 0, 1, '  Pairs')
        self.pairsbox = Text(pair_c, font=self.f_entry, bd=0, height=3, wrap=WORD,
                             highlightthickness=1, highlightbackground=BORDER,
                             highlightcolor=ACCENT, bg='#F8FAFC',
                             insertbackground=TEXT_PRIMARY)
        self.pairsbox.pack(fill=X, pady=(0, 6))

        pairsFromText = open('pairs.txt').read()
        self.pairsbox.insert(END, pairsFromText)

        self.browseButtonPairs = Button(pair_c, text=' Browse... ', font=self.f_btn,
                                        bg=ACCENT, fg='white', bd=0, padx=10, pady=2,
                                        activebackground=ACCENT_HOVER, activeforeground='white',
                                        cursor='hand2', command=self.browse_pairs,
                                        relief='flat', highlightthickness=0)
        self.browseButtonPairs.pack(anchor=W)
        _hover_bind(self.browseButtonPairs, ACCENT, ACCENT_HOVER)

        # ═══════════════════════════════════════════════════════
        #  ROW 4 — Console (expands, taller)
        # ═══════════════════════════════════════════════════════
        status_outer = Frame(body, bg=BG)
        status_outer.grid(row=4, column=0, columnspan=2, sticky='nsew', padx=3, pady=3)
        status_outer.columnconfigure(0, weight=1)
        status_outer.rowconfigure(1, weight=1)

        s_hdr = Frame(status_outer, bg=SECTION_HDR, highlightbackground=BORDER,
                      highlightthickness=1)
        s_hdr.grid(row=0, column=0, sticky='ew')
        Label(s_hdr, text='  Console Output', font=self.f_section, bg=SECTION_HDR,
              fg=TEXT_PRIMARY, anchor=W, padx=10, pady=3).pack(fill=X)

        status_inner = Frame(status_outer, bg=STATUS_BG, highlightbackground=BORDER,
                             highlightthickness=1)
        status_inner.grid(row=1, column=0, sticky='nsew')
        status_inner.columnconfigure(0, weight=1)
        status_inner.rowconfigure(0, weight=1)

        self.statusbar = ScrolledText(status_inner, state='disabled', font=self.f_status,
                                      bg=STATUS_BG, fg=STATUS_FG, bd=0,
                                      insertbackground=STATUS_FG,
                                      selectbackground='#334155', wrap=WORD,
                                      padx=10, pady=6, height=14)
        self.statusbar.grid(row=0, column=0, sticky='nsew')
        self.statusbar.tag_configure('center', justify=CENTER)

        # ═══════════════════════════════════════════════════════
        #  ROW 5 — Action buttons (same style as Browse)
        # ═══════════════════════════════════════════════════════
        btn_frame = Frame(body, bg=BG)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=(6, 4))

        self.runbutton = Button(btn_frame, text='  Run Analysis  ', font=self.f_btn_lg,
                                fg='white', bg=ACCENT, bd=0, padx=18, pady=5,
                                activebackground=ACCENT_HOVER, activeforeground='white',
                                cursor='hand2', command=self.runbutton_click,
                                relief='flat', highlightthickness=0)
        self.runbutton.pack(side=LEFT, padx=10)
        _hover_bind(self.runbutton, ACCENT, ACCENT_HOVER)

        self.openbutton = Button(btn_frame, text='  Open Result  ', font=self.f_btn_lg,
                                 fg='white', bg=ACCENT, bd=0, padx=18, pady=5,
                                 activebackground=ACCENT_HOVER, activeforeground='white',
                                 cursor='hand2', command=self.open_click,
                                 relief='flat', highlightthickness=0)
        self.openbutton.pack(side=LEFT, padx=10)
        self.openbutton.configure(state=DISABLED)
        _hover_bind(self.openbutton, ACCENT, ACCENT_HOVER)

        # ── Startup message (centered) ──────────────────────────
        self._status_centered('>> mePROD App v3.0.0 Started! <<\n')
        self._status_centered('━' * 46 + '\n')

    def _status_centered(self, text):
        """Insert text centered in the console."""
        self.statusbar.configure(state='normal')
        self.statusbar.insert(END, text, 'center')
        self.statusbar.see(END)
        self.statusbar.configure(state='disabled')

    def Message(self, title, message):
        messagebox.showinfo(title=title, message=message)

    def update_status_box(self, text):
        self.statusbar.configure(state='normal')
        self.statusbar.insert(END, text)
        self.statusbar.see(END)
        self.statusbar.configure(state='disabled')

    def clear_status_box(self):
        self.statusbar.configure(state='normal')
        self.statusbar.delete(1.0, END)
        self.statusbar.see(END)
        self.statusbar.configure(state='disabled')

    def check_main_thread(self):
        self.root.update()
        if self.myThread.is_alive():
            self.root.after(1000, self.check_main_thread)
        else:
            self.x = True

    def open_click(self):
        os.startfile(f'{self.outputLocationPath}/{self.outputLocation.strip()}.xlsx')

    def browse(self):
        self.filename = filedialog.askopenfile(parent=self.root, mode='rb', title='Choose a PSMs file')
        self.filenamePretify = str(self.filename).split('/')[-1].split("'>")[0]
        if self.filenamePretify == "None":
            self.Message('Error!', 'Please choose a file!')
            return 0
        self.fileLabel.configure(text=self.filenamePretify, fg=TEXT_PRIMARY,
                                 font=self.f_label)
        self.update_status_box(f'  File selected: "{self.filenamePretify}"\n')
        self.outputLocationPath = str(self.filename).split("'")[1].replace(
            str(self.filename).split("'")[1].split("/")[-1], '')

    def browse_condition(self):
        self.filename_condition = filedialog.askopenfile(parent=self.root, mode='rb',
                                                         title='Please, choose a condition text file.')
        self.filenamePretify_condition = str(self.filename_condition).split('/')[-1].split("'>")[0]
        if self.filenamePretify_condition == "None":
            self.Message('Error!', 'Please choose a file!')
            return 0
        self.outputLocationPath_condition = str(self.filename_condition).split("'")[1].replace(
            str(self.filename_condition).split("'")[1].split("/")[-1], '')
        condtionsFromText = str(
            open(self.outputLocationPath_condition + self.filenamePretify_condition).read()).strip()
        self.conditionbox.delete('1.0', END)
        self.conditionbox.insert(END, condtionsFromText)

    def browse_pairs(self):
        self.filename_pairs = filedialog.askopenfile(parent=self.root, mode='rb',
                                                      title='Please, choose a pairs text file.')
        self.filenamePretify_pairs = str(self.filename_pairs).split('/')[-1].split("'>")[0]
        if self.filenamePretify_pairs == "None":
            self.Message('Error!', 'Please choose a file!')
            return 0
        self.outputLocationPath_pairs = str(self.filename_pairs).split("'")[1].replace(
            str(self.filename_pairs).split("'")[1].split("/")[-1], '')
        pairsFromText = str(
            open(self.outputLocationPath_pairs + self.filenamePretify_pairs).read()).strip()
        self.pairsbox.delete('1.0', END)
        self.pairsbox.insert(END, pairsFromText)

    def runbutton_click(self):
        self.runbutton.configure(state=DISABLED)
        self.myThread = Thread(target=self.engine)
        self.myThread.daemon = True
        self.myThread.start()
        self.root.after(1000, self.check_main_thread)

    def reportAndExport(self, details, data, outputLocation):
        wb = Workbook()
        ws = wb.active
        ws.title = "Info"

        # Program title
        program_title = f"mePROD App {__version__} by S. Bozkurt @2026"

        ws['H1'] = program_title
        ws['H1'].font = sty.Font(size=18, bold=True, color="0072BB")
        ws['H1'].alignment = sty.Alignment(horizontal="center", vertical="center")

        # Splitting the details between columns H and I
        details_split = [(key, value) for key, value in details.items()]

        for index, (label, value) in enumerate(details_split, start=3):
            ws.cell(row=index, column=8, value=label).alignment = sty.Alignment(horizontal="right")
            ws.cell(row=index, column=8).font = sty.Font(size=16)
            ws.cell(row=index, column=9, value=value).alignment = sty.Alignment(horizontal="left")
            ws.cell(row=index, column=9).font = sty.Font(size=16)

        current_date = datetime.now().strftime('%Y-%m-%d')

        ws["H18"] = "Date:"
        ws["H18"].alignment = sty.Alignment(horizontal="right")
        ws["H18"].font = sty.Font(size=16)
        ws["I18"] = current_date
        ws["I18"].font = sty.Font(size=16)

        ws["H19"] = "Processed by:"
        ws["H19"].alignment = sty.Alignment(horizontal="right")
        ws["H19"].font = sty.Font(size=16)
        ws["I19"] = f"User {os.getlogin()}"
        ws["I19"].font = sty.Font(size=16)

        # Add results to a new sheet
        ws_results = wb.create_sheet("Results")

        header_font = sty.Font(bold=True, size=13)
        center_alignment = sty.Alignment(horizontal='center', vertical='center')

        for c_idx, col_name in enumerate(data.columns, 1):
            cell = ws_results.cell(row=1, column=c_idx, value=col_name)
            cell.font = header_font
            cell.alignment = center_alignment

        for r_idx, row in enumerate(data.iterrows(), 2):
            for c_idx, value in enumerate(row[1], 1):
                ws_results.cell(row=r_idx, column=c_idx, value=value)

        wb.save(outputLocation)

        return outputLocation

    def engine(self):
        try:
            if '.xlsx' in self.filenamePretify:
                self.fileRead = pd.read_excel(self.outputLocationPath + self.filenamePretify)
            elif '.txt' in self.filenamePretify:
                self.fileRead = pd.read_csv(self.outputLocationPath + self.filenamePretify, sep='\t', header=0)
            self.update_status_box('  Reading file...\n')
        except Exception:
            self.update_status_box('  ERROR: Please choose a file before running!\n')
            self.Message('Error!', 'Please choose a file before run!')
            self.runbutton.configure(state=NORMAL)
            return 0

        self.update_status_box('  File loaded successfully.\n')

        # Get MS level
        ms_level = self.msLevelVar.get()

        # Conditions
        self.conditions = self.conditionbox.get("1.0", END)

        condtionsFile = open(f'{self.outputLocationPath}/condtions.txt', 'w')
        condtionsFile.write(self.conditions)
        condtionsFile.close()

        condtionsFile = open(f'condtions.txt', 'w')
        condtionsFile.write(self.conditions)
        condtionsFile.close()

        conditionsFinal = self.conditions.split(',')
        conditionsFinal[-1] = conditionsFinal[-1].strip()

        # Pairs
        self.pairs = self.pairsbox.get("1.0", END)

        pairsFile = open(f'{self.outputLocationPath}/pairs.txt', 'w')
        pairsFile.write(self.pairs)
        pairsFile.close()

        pairsFile = open(f'pairs.txt', 'w')
        pairsFile.write(self.pairs)
        pairsFile.close()

        pairsFinal = self.pairs.split(';')
        pairsFinal[-1] = pairsFinal[-1].strip()
        pairsFinal = [i.strip() for i in pairsFinal]
        pairsFinal = [i.lstrip() for i in pairsFinal]
        pairsFinal = [pairs.split('/') for pairs in pairsFinal]

        normalization_type = self.normVar.get()
        finalStatisticalMethod = self.statisticVar.get().strip()

        if pairsFinal == [['']]:
            finalStatisticalMethod = None
            pairsFinalOutput = 'None'
        else:
            pairsFinalOutput = self.pairs.strip()

        self.update_status_box(f'  MS Level     : {ms_level}\n')
        self.update_status_box(f'  Conditions   : {self.conditions.strip()}\n')
        self.update_status_box(f'  Pairs        : {pairsFinalOutput}\n')
        self.update_status_box(f'  Normalization: {normalization_type.strip()}\n')
        self.update_status_box(f'  Statistics   : {str(finalStatisticalMethod).strip()}\n')
        self.update_status_box('  ' + '─' * 56 + '\n')
        self.update_status_box('  Running...\n')

        try:
            randomReportName = f'reports_{str(random.randint(1, 100000))}'
            mePROD_class = mePROD(self.outputLocationPath, randomReportName)
            self.data = mePROD_class.engine(
                self.fileRead, conditionsFinal, pairsFinal,
                normalization_type, finalStatisticalMethod, ms_level
            )
        except Exception as e:
            self.runbutton.configure(state=NORMAL)
            self.update_status_box(f'  ERROR: {e}\n')
            self.Message('Error!',
                         f'An error occurred: {e}\nPlease fix and rerun or contact developer via {__email__}!')
            return 0

        try:
            if self.data == 0:
                self.update_status_box('  ERROR: Baseline channel not found!\n')
                self.Message('Error!', 'Please provide light/baseline channel!')
                self.runbutton.configure(state=NORMAL)
                return 0
        except Exception:
            pass

        self.update_status_box('  Processing complete.\n')
        self.update_status_box('  Saving data...\n')

        self.outputLocation = self.outputNamebox.get("1.0", END)

        try:
            self.data.to_excel(f'{self.outputLocationPath}/{self.outputLocation.strip()}.xlsx',
                               index=False, engine="openpyxl")

            self.datamePROD = pd.read_excel(
                f'{self.outputLocationPath}/{self.outputLocation.strip()}.xlsx')

            # Assign gene names
            self.data = mePROD_class.GeneNameEngine(self.datamePROD)

            # Determine mitochondrial proteins
            self.data = mePROD_class.mito_human(self.data)

            # Assign significance markers
            self.data = mePROD_class.significantAssig(self.data)

            # Read report file
            file_path = os.path.join(self.outputLocationPath, f"{randomReportName}.txt")

            with open(file_path, "r") as file:
                content = file.read()

            totalPeptides = re.search(r"The number of total peptides: (\d+)", content).group(1)
            heavyPeptides = re.search(r"The number of heavy peptides: (\d+)", content).group(1)
            mitosHeavyPeptides = re.search(r"The number of mitochondrial heavy peptides: (\d+)", content).group(1)
            HeavyProteins = re.search(r"The number of heavy proteins: (\d+)", content).group(1)
            mitosHeavyProteins = re.search(r"The number of mitochondrial heavy proteins: (\d+)", content).group(1)

            os.remove(file_path)

            details = {
                "Version of the program:": f"{__version__}",
                "MS Level:": ms_level,
                "The number of total peptides:": totalPeptides,
                "The number of heavy peptides:": heavyPeptides,
                "The number of mitochondrial heavy peptides:": mitosHeavyPeptides,
                "The number of heavy proteins:": HeavyProteins,
                "The number of mitochondrial heavy proteins:": mitosHeavyProteins,
                "": "",
                "Input file:": self.filenamePretify.strip(),
                "Conditions:": self.conditions.strip(),
                "Pairs:": pairsFinalOutput,
                "Normalization:": normalization_type.strip(),
                "Statistics:": str(finalStatisticalMethod)
            }

            self.reportAndExport(details, self.data,
                                 f'{self.outputLocationPath}/{self.outputLocation.strip()}.xlsx')

            self.update_status_box(f'  Saved as {self.outputLocation.strip()}.xlsx\n')
            self.update_status_box('  ' + '─' * 56 + '\n')
            self.Message('Finished!', 'Analysis completed successfully!')
            self.openbutton.configure(state=NORMAL)
            self.runbutton.configure(state=NORMAL)
        except Exception as e:
            self.update_status_box(f'  ERROR: {e}\n')
            self.Message('Error!',
                         f'An error occurred: {e}\nPlease fix and rerun or contact developer via {__email__}!')
            self.runbutton.configure(state=NORMAL)
